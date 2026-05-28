# ragapp/machine/table_machine.py
from __future__ import annotations

import os
import re
import csv
import json
import math
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET, require_http_methods
from django.contrib.admin.views.decorators import staff_member_required
from django.utils import timezone

from .feature_config import (
    PUBLIC_MAX_CSV_ROWS,
    MEDIA_ROOT,
    TABLE_DATA_DIR,
    _int,
)
from .media_helpers import _log

log = logging.getLogger(__name__)

# 선택: 표 스키마 / 검색 규칙 모델 (없을 수도 있음)
try:
    from ragapp.models import TableSchema, TableSearchRule  # type: ignore
except Exception:  # pragma: no cover
    TableSchema = None  # type: ignore
    TableSearchRule = None  # type: ignore

# Vertex 임베딩/LLM 헬퍼
try:
    from ragapp.services.vertex_embed import (
        embed_texts_vertex as embed_texts,
        infer_table_query_with_vertex,
    )
except Exception:  # pragma: no cover
    from ragapp.services.vertex_embed import embed_texts_vertex as embed_texts
    infer_table_query_with_vertex = None  # type: ignore

# Chroma(표)
CHROMA_TABLE_IMPORT_ERR = None
try:
    from ragapp.services import chroma_media as cm

    add_table_rows = cm.add_table_rows
    search_table_by_text_embedding = cm.search_table_by_text_embedding
except Exception as e:  # pragma: no cover
    CHROMA_TABLE_IMPORT_ERR = f"{e.__class__.__name__}: {e}"
    add_table_rows = None  # type: ignore
    search_table_by_text_embedding = None  # type: ignore


def _as_path(p: Any) -> Path:
    try:
        return p if isinstance(p, Path) else Path(str(p))
    except Exception:
        return Path(str(p or "."))


def _is_safe_table_name(name: str) -> bool:
    """
    어드민 전용이라도 파일 경로로 쓰이므로 최소한의 경로 이탈 방지.
    - /, \, .. 금지
    """
    if not name:
        return False
    if "/" in name or "\\" in name:
        return False
    if ".." in name:
        return False
    return True


# ────────────────────────────────────────────────
# (3) 표 인덱싱 (CSV + 엑셀)  ✅ 어드민(스태프) 전용
# ────────────────────────────────────────────────
@staff_member_required
@never_cache
@require_http_methods(["GET", "POST"])
def table_index_view(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": True,  # ✅ 스태프는 항상 업로드 가능
                "max_rows": PUBLIC_MAX_CSV_ROWS,
            },
        )

    # POST
    if add_table_rows is None:
        msg = "표 인덱싱 모듈을 불러올 수 없습니다(chroma_media import 실패)."
        try:
            u = getattr(request, "user", None)
            if u and getattr(u, "is_staff", False) and CHROMA_TABLE_IMPORT_ERR:
                msg = f"표 인덱싱 모듈 import 실패: {CHROMA_TABLE_IMPORT_ERR}"
        except Exception:
            pass
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": True,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "error": msg,
            },
            status=500,
        )

    table_name = (request.POST.get("table_name") or "").strip()
    f = request.FILES.get("csvfile")
    limit = _int(request.POST.get("limit"), 0)

    if not table_name or not f:
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": True,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "error": "표 이름과 파일을 모두 입력해 주세요.",
            },
        )

    if not _is_safe_table_name(table_name):
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": True,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "error": "표 이름에 사용할 수 없는 문자가 포함되어 있습니다. (/, \\, .. 금지)",
            },
        )

    media_root = _as_path(MEDIA_ROOT).resolve()
    table_data_dir = _as_path(TABLE_DATA_DIR).resolve()

    root = media_root / "tables" / timezone.now().strftime("%Y/%m")
    try:
        root.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    safe_name = os.path.basename(getattr(f, "name", "table.csv") or "table.csv")
    dst = root / f"{timezone.now().strftime('%Y%m%d%H%M%S%f')}_{safe_name}"

    rows: List[Dict[str, Any]] = []

    try:
        with open(dst, "wb") as out:
            for chunk in f.chunks():
                out.write(chunk)

        ext = dst.suffix.lower()

        if ext in [".xlsx", ".xls"]:
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError(
                    "엑셀 파일(.xlsx, .xls)을 읽으려면 'openpyxl' 패키지가 필요합니다. "
                    "터미널에서 'pip install openpyxl' 실행 후 다시 시도해 주세요."
                )

            wb = openpyxl.load_workbook(dst, data_only=True)
            sheet = wb.active

            header = None
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(c).strip() if c is not None else "" for c in row]
                break

            if not header or all(not h for h in header):
                raise RuntimeError("엑셀 파일에서 열 이름(첫 줄)을 찾을 수 없습니다.")

            max_rows = int(PUBLIC_MAX_CSV_ROWS)
            limit_rows = min(limit or max_rows, max_rows)

            for _idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                rec: Dict[str, Any] = {}
                for i, col in enumerate(header):
                    col_name = col or f"col_{i+1}"
                    val = row[i] if row and i < len(row) else None
                    rec[col_name] = val
                rows.append(rec)
                if limit_rows and len(rows) >= limit_rows:
                    break

        else:
            with open(dst, "r", encoding="utf-8", newline="") as rf:
                reader = csv.DictReader(rf)
                max_rows = int(PUBLIC_MAX_CSV_ROWS)
                limit_rows = min(limit or max_rows, max_rows)
                for row in reader:
                    rows.append(row)
                    if limit_rows and len(rows) >= limit_rows:
                        break

        if not rows:
            return render(
                request,
                "ragapp/table_index.html",
                {
                    "allow_upload": True,
                    "max_rows": PUBLIC_MAX_CSV_ROWS,
                    "error": "표 안에 읽을 수 있는 줄이 없습니다.",
                },
            )

        # 스키마 저장 (있으면)
        try:
            if TableSchema is not None:
                cols = list(rows[0].keys())
                sample_rows = rows[:5]

                def _infer_type(val: Any) -> str:
                    if val is None:
                        return "text"
                    s = str(val).strip()
                    if not s:
                        return "text"
                    try:
                        float(str(s).replace(",", ""))
                        return "number"
                    except Exception:
                        pass
                    try:
                        datetime.fromisoformat(s)
                        return "date"
                    except Exception:
                        return "text"

                column_types: Dict[str, str] = {}
                for col in cols:
                    inferred = "text"
                    for r in sample_rows:
                        v = r.get(col)
                        if v not in (None, ""):
                            inferred = _infer_type(v)
                            break
                    column_types[col] = inferred

                TableSchema.objects.update_or_create(
                    table_name=table_name,
                    defaults={
                        "columns": cols,
                        "column_types": column_types,
                        "sample_rows": sample_rows,
                    },
                )
        except Exception:
            pass

        # 원본 JSON 저장
        try:
            try:
                table_data_dir.mkdir(parents=True, exist_ok=True)
            except Exception:
                pass
            json_path = table_data_dir / f"{table_name}.json"
            with open(json_path, "w", encoding="utf-8") as jf:
                json.dump(rows, jf, ensure_ascii=False)
        except Exception:
            pass

        # 벡터 인덱싱
        def row_to_str(r: Dict[str, Any]) -> str:
            return " | ".join(f"{k}:{r.get(k,'')}" for k in r.keys())

        texts = [row_to_str(r) for r in rows]
        if not texts:
            raise RuntimeError("인덱싱할 행이 없습니다.")

        embs = embed_texts(texts)
        added = add_table_rows(
            table_name=table_name,
            rows=rows,
            embeddings=embs,
        )

        _log(
            request,
            "table_index",
            table_name,
            True,
            {"rows": len(rows), "added": added},
        )

        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": True,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "ok": True,
                "table_name": table_name,
                "added": added,
                "filename": dst.name,
            },
        )

    except Exception as e:
        rows_count = len(rows) if isinstance(rows, list) else 0
        _log(
            request,
            "table_index",
            table_name,
            False,
            {"rows": rows_count, "error": str(e)},
        )
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": True,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "error": str(e),
            },
        )


# ────────────────────────────────────────────────
# (4) 표 검색 + 그룹/집계  ✅ 어드민(스태프) 전용 / quota 제거
# ────────────────────────────────────────────────
ALLOWED_AGG = {"", "count", "sum", "avg", "min", "max"}

AGG_HINTS: dict[str, list[str]] = {
    "sum": ["합계", "총 ", "전체", "총액", "총매출", "total"],
    "avg": ["평균", "평균적으로", "average", "avg"],
    "max": ["최대", "가장 큰", "제일 큰", "가장 높은", "top"],
    "min": ["최소", "가장 작은", "가장 낮은"],
    "count": ["개수", "건수", "몇 개", "몇개", "몇 명", "몇명", "row 수"],
}

COLUMN_SYNONYMS: dict[str, list[str]] = {
    "region": ["지역", "지역별", "도시", "시도", "branch", "지점"],
    "product": ["상품", "메뉴", "제품", "메뉴명", "item"],
    "channel": ["채널", "판매채널", "판매 경로", "sales channel"],
    "date": ["날짜", "일자", "date", "일별"],
    "sales": ["매출", "매출액", "금액", "판매금액", "revenue", "sales"],
}

NUMERIC_HINTS = ["sales", "amount", "revenue", "price", "qty", "quantity", "count"]


def _safe_json_dict(text: Any) -> Dict[str, Any]:
    if not text:
        return {}
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def _safe_json_list(text: Any) -> List[Any]:
    if not text:
        return []
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except Exception:
        pass
    return []


def _load_table_search_config(
    table: str,
) -> Tuple[Dict[str, List[str]], Dict[str, List[str]], List[str], float, bool]:
    agg_hints = {k: list(v) for k, v in AGG_HINTS.items()}
    column_synonyms = {k: list(v) for k, v in COLUMN_SYNONYMS.items()}
    numeric_hints = list(NUMERIC_HINTS)
    min_sim = 0.35
    hard_filter_enabled = True

    if TableSearchRule is None:
        return agg_hints, column_synonyms, numeric_hints, min_sim, hard_filter_enabled

    try:
        qs = TableSearchRule.objects.filter(is_active=True)
        rule = None
        if table:
            rule = qs.filter(table_name=table).order_by("-updated_at", "-id").first()
        if rule is None:
            rule = qs.filter(table_name__in=["", None]).order_by("-updated_at", "-id").first()
    except Exception:
        rule = None

    if not rule:
        return agg_hints, column_synonyms, numeric_hints, min_sim, hard_filter_enabled

    try:
        if getattr(rule, "min_sim", None) is not None:
            min_sim = float(rule.min_sim)
    except Exception:
        pass

    hard_filter_enabled = bool(getattr(rule, "hard_filter_enabled", True))

    override_agg = _safe_json_dict(getattr(rule, "agg_hints_json", None))
    for key, words in override_agg.items():
        if isinstance(words, list):
            agg_hints[str(key)] = [str(w) for w in words]
        elif isinstance(words, str):
            agg_hints[str(key)] = [words]

    override_syn = _safe_json_dict(getattr(rule, "column_synonyms_json", None))
    for key, syns in override_syn.items():
        if isinstance(syns, list):
            column_synonyms[str(key)] = [str(s) for s in syns]
        elif isinstance(syns, str):
            column_synonyms[str(key)] = [syns]

    override_num = _safe_json_list(getattr(rule, "numeric_hints_json", None))
    if override_num:
        numeric_hints = [str(x) for x in override_num]

    return agg_hints, column_synonyms, numeric_hints, min_sim, hard_filter_enabled


def _get_table_schema_info(table_name: str) -> tuple[list[str], dict[str, str]]:
    if TableSchema is None or not table_name:
        return [], {}

    try:
        schema = (
            TableSchema.objects.filter(table_name=table_name)
            .order_by("-updated_at", "-created_at", "-id")
            .first()
        )
    except Exception:
        schema = None

    if schema is None:
        return [], {}

    cols_raw = getattr(schema, "columns", None)
    col_types_raw = getattr(schema, "column_types", None) or getattr(schema, "column_types_json", None)

    cols: list[str] = []
    if isinstance(cols_raw, list):
        for item in cols_raw:
            if isinstance(item, dict):
                name = item.get("name") or item.get("column") or item.get("key")
            else:
                name = str(item)
            if name and name not in cols:
                cols.append(name)
    elif isinstance(cols_raw, dict):
        cols = list(cols_raw.keys())
    elif isinstance(cols_raw, str):
        try:
            j = json.loads(cols_raw)
            if isinstance(j, list):
                for item in j:
                    if isinstance(item, dict):
                        name = item.get("name") or item.get("column") or item.get("key")
                    else:
                        name = str(item)
                    if name and name not in cols:
                        cols.append(name)
            elif isinstance(j, dict):
                cols = list(j.keys())
        except Exception:
            cols = [c.strip() for c in cols_raw.split(",") if c.strip()]

    col_types: dict[str, str] = {}
    if isinstance(col_types_raw, dict):
        col_types = {str(k): str(v) for k, v in col_types_raw.items()}
    elif isinstance(col_types_raw, str):
        try:
            j = json.loads(col_types_raw)
            if isinstance(j, dict):
                col_types = {str(k): str(v) for k, v in j.items()}
        except Exception:
            pass

    return cols, col_types


def _guess_agg_from_question(q: str, agg_hints: Dict[str, List[str]]) -> str:
    q_lower = q.lower()
    for agg_key, words in agg_hints.items():
        for w in words:
            if w in q or w in q_lower:
                return agg_key
    return ""


def _auto_fill_table_and_agg(
    q: str,
    table: str,
    group_by: str,
    agg_field: str,
    agg: str,
    agg_hints: Dict[str, List[str]],
    column_synonyms: Dict[str, List[str]],
    numeric_hints: List[str],
) -> tuple[str, str, str, str]:
    if (not table) and TableSchema is not None:
        try:
            distinct_tables = list(
                TableSchema.objects.order_by("table_name")
                .values_list("table_name", flat=True)
                .distinct()
            )
        except Exception:
            distinct_tables = []
        if len(distinct_tables) == 1:
            table = distinct_tables[0]

    if TableSchema is None or not table:
        return table, group_by, agg_field, agg

    cols, col_types = _get_table_schema_info(table)
    if not cols:
        return table, group_by, agg_field, agg

    if not agg:
        agg = _guess_agg_from_question(q, agg_hints)

    if not agg_field and agg:
        numeric_cols = [c for c in cols if col_types.get(c) in ("number", "numeric", "float", "int")]
        chosen = ""
        q_lower = q.lower()

        for c in numeric_cols:
            name_lower = c.lower()
            if any(h in name_lower for h in numeric_hints):
                chosen = c
                break
            if name_lower in q_lower:
                chosen = c
                break

        if not chosen and len(numeric_cols) == 1:
            chosen = numeric_cols[0]

        if chosen:
            agg_field = chosen

    if not group_by and agg and agg_field:
        q_lower = q.lower()
        text_cols = [c for c in cols if col_types.get(c, "text") not in ("number", "numeric", "float", "int")]

        candidate_scores: list[tuple[int, str]] = []
        for c in text_cols:
            score = 0
            name_lower = c.lower()

            if name_lower in q_lower:
                score += 5

            for key, syns in column_synonyms.items():
                if key in name_lower:
                    if any(s in q for s in syns):
                        score += 4
                    else:
                        score += 1

            if score > 0:
                candidate_scores.append((score, c))

        if candidate_scores:
            candidate_scores.sort(reverse=True)
            group_by = candidate_scores[0][1]
        else:
            if len(text_cols) == 1:
                group_by = text_cols[0]

    return table, group_by, agg_field, agg


def _infer_columns(hit_rows: List[Dict[str, Any]], table: Optional[str]) -> List[str]:
    if not hit_rows:
        return []

    seen: List[str] = []
    for r in hit_rows:
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in seen:
                seen.append(k)

    if not table and "_table" in seen:
        seen.remove("_table")
        seen.insert(0, "_table")

    return seen


def _to_float(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(",", "").strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _apply_group_agg(
    hit_rows: List[Dict[str, Any]],
    group_by: str,
    agg: str,
    agg_field: str,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    groups: Dict[str, List[float]] = {}
    counts: Dict[str, int] = {}

    for r in hit_rows:
        if not isinstance(r, dict):
            continue

        key = str(r.get(group_by, "") or "(값 없음)")
        counts[key] = counts.get(key, 0) + 1

        if agg == "count" or not agg_field:
            continue

        val = _to_float(r.get(agg_field))
        if val is None:
            continue
        groups.setdefault(key, []).append(val)

    rows_out: List[Dict[str, Any]] = []

    if agg == "count" or not agg_field:
        columns = [group_by, "rows"]
        for key, count in counts.items():
            rows_out.append({group_by: key, "rows": count})
    else:
        col_name = f"{agg}_{agg_field}"
        columns = [group_by, "rows", col_name]
        for key, nums in groups.items():
            if not nums:
                continue

            if agg == "sum":
                value = sum(nums)
            elif agg == "avg":
                value = sum(nums) / len(nums)
            elif agg == "min":
                value = min(nums)
            elif agg == "max":
                value = max(nums)
            else:
                value = len(nums)

            rows_out.append({group_by: key, "rows": counts.get(key, len(nums)), col_name: value})

    rows_out.sort(key=lambda r: str(r.get(group_by, "")))
    return rows_out, columns


def _load_table_rows_from_file(table: str) -> List[Dict[str, Any]]:
    try:
        if not table:
            return []
        base = _as_path(TABLE_DATA_DIR).resolve()
        path = base / f"{table}.json"
        if not path.exists():
            return []
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [r for r in data if isinstance(r, dict)]
        return []
    except Exception:
        return []


def _apply_filters(
    rows: List[Dict[str, Any]],
    filters: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    if not filters:
        return rows

    def _match(row: Dict[str, Any], flt: Dict[str, Any]) -> bool:
        col = flt.get("column") or flt.get("field")
        if not col:
            return True
        op = (flt.get("op") or "=").lower()
        val = flt.get("value")
        cell = row.get(col)
        if cell is None:
            return False

        s = str(cell)
        if op in ("=", "eq"):
            if isinstance(val, list):
                return s in [str(v) for v in val]
            return s == str(val)
        if op in ("contains", "like"):
            return str(val) in s
        if op == "in":
            if not isinstance(val, list):
                return False
            return s in [str(v) for v in val]

        try:
            cnum = float(str(cell).replace(",", ""))
            vnum = float(str(val).replace(",", ""))
        except Exception:
            return False

        if op in (">", "gt"):
            return cnum > vnum
        if op in (">=", "ge"):
            return cnum >= vnum
        if op in ("<", "lt"):
            return cnum < vnum
        if op in ("<=", "le"):
            return cnum <= vnum

        return True

    out: List[Dict[str, Any]] = []
    for r in rows:
        ok = True
        for f in filters:
            if not _match(r, f):
                ok = False
                break
        if ok:
            out.append(r)
    return out


def _hard_filter_rows_by_question(
    q: str,
    rows: List[Dict[str, Any]],
    columns: List[str],
) -> List[Dict[str, Any]]:
    if not q or not rows or not columns:
        return rows

    q_norm = q.replace(" ", "")
    candidates = rows

    for col in columns:
        if col == "_table":
            continue

        values = sorted({str(r.get(col, "")).strip() for r in rows if r.get(col) not in (None, "")})
        if not values:
            continue

        hit_vals = []
        for v in values:
            v_norm = v.replace(" ", "")
            if not v_norm:
                continue
            if v in q or v_norm in q_norm:
                hit_vals.append(v)

        if hit_vals:
            hit_set = set(hit_vals)
            new_candidates = [r for r in candidates if str(r.get(col, "")).strip() in hit_set]
            if new_candidates:
                candidates = new_candidates

    return candidates or rows


@staff_member_required
@require_GET
def table_search_view(request: HttpRequest) -> HttpResponse:
    q = (request.GET.get("q") or "").strip()

    def _to_int_param(name: str, default: int) -> int:
        try:
            v = int(request.GET.get(name, default))
            return max(1, v)
        except Exception:
            return default

    size = _to_int_param("size", 12)
    page = _to_int_param("page", 1)
    try:
        k = int(request.GET.get("k", 200))
    except Exception:
        k = 200

    table = (request.GET.get("table") or "").strip()
    group_by = (request.GET.get("group_by") or "").strip()
    agg_field = (request.GET.get("agg_field") or "").strip()
    agg = (request.GET.get("agg") or "").strip().lower()
    if agg not in ALLOWED_AGG:
        agg = ""

    table_names: list[str] = []
    if TableSchema is not None:
        try:
            table_names = list(
                TableSchema.objects.order_by("table_name")
                .values_list("table_name", flat=True)
                .distinct()
            )
        except Exception:
            table_names = []

    (
        agg_hints_cfg,
        column_synonyms_cfg,
        numeric_hints_cfg,
        min_sim,
        hard_filter_enabled,
    ) = _load_table_search_config(table)

    columns: list[str] = []
    rows: list[dict] = []
    total: int = 0
    page_count: int = 1
    error_msg: Optional[str] = None
    used_loose: bool = False

    if not q:
        ctx = {
            "q": q,
            "size": size,
            "page": page,
            "k": k,
            "table": table,
            "group_by": group_by,
            "agg_field": agg_field,
            "agg": agg,
            "columns": columns,
            "rows": rows,
            "total": total,
            "page_count": page_count,
            "error_msg": error_msg,
            "table_names": table_names,
            "used_loose": used_loose,
        }
        return render(request, "ragapp/table_search.html", ctx)

    orig_table = table
    table, group_by, agg_field, agg = _auto_fill_table_and_agg(
        q=q,
        table=table,
        group_by=group_by,
        agg_field=agg_field,
        agg=agg,
        agg_hints=agg_hints_cfg,
        column_synonyms=column_synonyms_cfg,
        numeric_hints=numeric_hints_cfg,
    )

    if not orig_table and table != orig_table:
        (
            agg_hints_cfg,
            column_synonyms_cfg,
            numeric_hints_cfg,
            min_sim,
            hard_filter_enabled,
        ) = _load_table_search_config(table)

    llm_plan: Optional[Dict[str, Any]] = None
    if infer_table_query_with_vertex is not None and TableSchema is not None:
        try:
            tables_for_llm: Dict[str, Dict[str, Any]] = {}
            for tname in table_names:
                cols, col_types = _get_table_schema_info(tname)
                sample_rows: List[Dict[str, Any]] = []
                try:
                    schema_obj = (
                        TableSchema.objects.filter(table_name=tname)
                        .order_by("-updated_at", "-created_at", "-id")
                        .first()
                    )
                    if schema_obj is not None:
                        sr = getattr(schema_obj, "sample_rows", None)
                        if isinstance(sr, list):
                            sample_rows = [r for r in sr if isinstance(r, dict)]
                        elif isinstance(sr, str):
                            try:
                                j = json.loads(sr)
                                if isinstance(j, list):
                                    sample_rows = [r for r in j if isinstance(r, dict)]
                            except Exception:
                                sample_rows = []
                except Exception:
                    sample_rows = []

                tables_for_llm[tname] = {
                    "columns": cols,
                    "column_types": col_types,
                    "sample_rows": sample_rows,
                }

            llm_plan = infer_table_query_with_vertex(
                question=q,
                tables=tables_for_llm,
                default_table=table or None,
            )
        except Exception as e:
            log.exception("infer_table_query_with_vertex 실패: %s", e)
            llm_plan = None

    plan_filters: List[Dict[str, Any]] = []
    if isinstance(llm_plan, dict):
        plan_table = (llm_plan.get("table") or "").strip()
        if plan_table and not table:
            table = plan_table

        if not group_by and llm_plan.get("group_by"):
            group_by = str(llm_plan.get("group_by"))

        if not agg_field and llm_plan.get("agg_field"):
            agg_field = str(llm_plan.get("agg_field"))

        if not agg and llm_plan.get("agg"):
            agg_candidate = str(llm_plan.get("agg")).lower()
            if agg_candidate in ALLOWED_AGG:
                agg = agg_candidate

        pf = llm_plan.get("filters") or llm_plan.get("where") or []
        if isinstance(pf, list):
            plan_filters = [f for f in pf if isinstance(f, dict)]

    try:
        if search_table_by_text_embedding is None:
            raise RuntimeError("표 검색 모듈을 불러올 수 없습니다(chroma_media import 실패).")

        q_vecs = embed_texts([q]) or []
        if not q_vecs:
            raise RuntimeError("임베딩을 만들 수 없습니다.")
        qv = q_vecs[0]

        res = search_table_by_text_embedding(text_embedding=qv, k=k) or {}
        metas_raw = res.get("metadatas") or []
        dists_raw = res.get("distances") or []

        if isinstance(metas_raw, list) and metas_raw:
            metas = metas_raw[0] if isinstance(metas_raw[0], list) else metas_raw
        else:
            metas = []

        if isinstance(dists_raw, list) and dists_raw:
            dists = dists_raw[0] if isinstance(dists_raw[0], list) else dists_raw
        else:
            dists = []

        if len(dists) < len(metas):
            dists = dists + [None] * (len(metas) - len(dists))
        elif len(dists) > len(metas):
            dists = dists[: len(metas)]

        strict_all: List[Dict[str, Any]] = []
        strict_filtered: List[Dict[str, Any]] = []
        loose_all: List[Dict[str, Any]] = []
        loose_filtered: List[Dict[str, Any]] = []

        MIN_SIM = float(min_sim or 0.0)

        for meta, dist in zip(metas, dists):
            if not isinstance(meta, dict):
                continue

            meta_table = (meta.get("table") or meta.get("table_name") or "").strip()
            row_json = meta.get("row_json") or meta.get("row") or meta.get("data") or {}

            if isinstance(row_json, str):
                try:
                    row = json.loads(row_json)
                except Exception:
                    row = {}
            elif isinstance(row_json, dict):
                row = row_json
            else:
                row = {}

            if not isinstance(row, dict) or not row:
                continue

            row_with_table = dict(row)
            if meta_table:
                row_with_table["_table"] = meta_table

            try:
                score = 1.0 - float(dist) if dist is not None else None
            except Exception:
                score = None

            match_table = (not table) or (not meta_table) or (meta_table == table)

            loose_all.append(row_with_table)
            if match_table:
                loose_filtered.append(row_with_table)

            if (score is None) or (score >= MIN_SIM):
                strict_all.append(row_with_table)
                if match_table:
                    strict_filtered.append(row_with_table)

        if strict_filtered:
            parsed_rows = strict_filtered
        elif strict_all:
            parsed_rows = strict_all
        elif loose_filtered:
            parsed_rows = loose_filtered
            used_loose = True
        else:
            parsed_rows = loose_all
            used_loose = True

        if not parsed_rows:
            if table:
                all_rows = _load_table_rows_from_file(table)
                if all_rows:

                    def _row_to_str(r: Dict[str, Any]) -> str:
                        return " ".join(f"{k}:{r.get(k,'')}" for k in r.keys())

                    keywords = [w for w in re.split(r"\s+", q) if w]
                    fallback_hits: List[Dict[str, Any]] = []
                    for r in all_rows:
                        s = _row_to_str(r)
                        if all(kw in s for kw in keywords):
                            fallback_hits.append(r)

                    parsed_rows = fallback_hits or all_rows
                    used_loose = True

            if not parsed_rows:
                ctx = {
                    "q": q,
                    "size": size,
                    "page": page,
                    "k": k,
                    "table": table,
                    "group_by": group_by,
                    "agg_field": agg_field,
                    "agg": agg,
                    "columns": [],
                    "rows": [],
                    "total": 0,
                    "page_count": 1,
                    "error_msg": None,
                    "table_names": table_names,
                    "used_loose": used_loose,
                }
                return render(request, "ragapp/table_search.html", ctx)

        if plan_filters:
            filtered = _apply_filters(parsed_rows, plan_filters)
            if filtered:
                parsed_rows = filtered

        col_order: Optional[List[str]] = None
        if TableSchema is not None and table:
            cols, _ = _get_table_schema_info(table)
            if cols:
                col_order = cols

        if col_order is None:
            col_order = _infer_columns(parsed_rows, table or None)

        if hard_filter_enabled:
            parsed_rows = _hard_filter_rows_by_question(q, parsed_rows, col_order or [])

        if group_by and agg and agg_field:
            rows_all, columns = _apply_group_agg(parsed_rows, group_by, agg, agg_field)
        else:
            rows_all = parsed_rows
            columns = col_order or []

        total = len(rows_all)
        page_count = max(1, math.ceil(total / max(1, size)))
        if page > page_count:
            page = page_count
        start = max(0, (page - 1) * size)
        end = start + size
        rows = rows_all[start:end]

        if total == 0 and table:
            all_rows = _load_table_rows_from_file(table)
            if all_rows:
                total = len(all_rows)
                page_count = max(1, math.ceil(total / max(1, size)))
                if page > page_count:
                    page = page_count
                start = max(0, (page - 1) * size)
                end = start + size
                rows = all_rows[start:end]
                columns = _infer_columns(all_rows, table)
                used_loose = True

    except Exception as e:
        error_msg = f"{e.__class__.__name__}: {e}"

    ctx = {
        "q": q,
        "size": size,
        "page": page,
        "k": k,
        "table": table,
        "group_by": group_by,
        "agg_field": agg_field,
        "agg": agg,
        "columns": columns,
        "rows": rows,
        "total": total,
        "page_count": page_count,
        "error_msg": error_msg,
        "table_names": table_names,
        "used_loose": used_loose,
    }
    return render(request, "ragapp/table_search.html", ctx)
