# ragapp/feature_views.py
from __future__ import annotations

import os
import csv
import mimetypes
import hashlib
import json
import re
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import logging

from django.conf import settings
from django.http import HttpRequest, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from django.views.decorators.cache import never_cache
from django.views.decorators.http import require_GET

from ragapp.models import MyLog

log = logging.getLogger(__name__)

# 선택: 표 스키마 / 검색 규칙 모델 (없을 수도 있음)
try:
    from ragapp.models import TableSchema, TableSearchRule  # type: ignore
except Exception:  # pragma: no cover
    TableSchema = None  # type: ignore
    TableSearchRule = None  # type: ignore

# Vertex 임베딩/LLM 헬퍼
try:
    from ragapp.services.vertex_embed import (
        embed_image_file,           # 이미지 → 벡터 (Vertex 멀티모달)
        embed_text_mm,              # 텍스트 → 멀티모달(이미지 검색용)
        embed_texts_vertex as embed_texts,  # 텍스트 → Vertex 텍스트 임베딩
        infer_table_query_with_vertex,      # (선택) 표 질의 해석용 LLM 헬퍼
    )
except Exception:  # pragma: no cover
    # infer_table_query_with_vertex 가 없거나 import 실패해도 나머지는 사용 가능하게
    from ragapp.services.vertex_embed import (
        embed_image_file,
        embed_text_mm,
        embed_texts_vertex as embed_texts,
    )
    infer_table_query_with_vertex = None  # type: ignore

# Chroma 벡터 스토어 (이미지/표 모두 여기에 저장)
from ragapp.services.chroma_media import (
    add_image_item,                  # media_images에 add
    search_images_by_text_embedding, # 텍스트벡터로 이미지 검색
    add_table_rows,                  # table_rows에 행 추가
    search_table_by_text_embedding,  # 표(행) 검색
)

# ────────────────────────────────────────────────
# 환경 스위치
# ────────────────────────────────────────────────
PUBLIC_ALLOW_UPLOAD_IMAGES = (os.environ.get("PUBLIC_ALLOW_UPLOAD_IMAGES", "1").lower() not in ("0", "false", "no"))
PUBLIC_ALLOW_UPLOAD_CSV    = (os.environ.get("PUBLIC_ALLOW_UPLOAD_CSV", "1").lower() not in ("0", "false", "no"))
PUBLIC_MAX_FILES           = int(os.environ.get("PUBLIC_MAX_FILES", "10"))
PUBLIC_MAX_FILE_MB         = int(os.environ.get("PUBLIC_MAX_FILE_MB", "15"))
PUBLIC_MAX_CSV_ROWS        = int(os.environ.get("PUBLIC_MAX_CSV_ROWS", "1000"))

CHROMA_MEDIA_DIR           = os.environ.get("CHROMA_MEDIA_DIR", "chroma_media")

# ✅ settings.py 에서 MEDIA_ROOT=/.../uploads, MEDIA_URL=/uploads/ 를 쓰는 걸 권장
MEDIA_ROOT = Path(getattr(settings, "MEDIA_ROOT", Path(settings.BASE_DIR) / "uploads")).resolve()
MEDIA_URL  = getattr(settings, "MEDIA_URL", "/uploads/")
MEDIA_ROOT.mkdir(parents=True, exist_ok=True)

# 표 원본 데이터를 JSON 으로 보관할 디렉터리
TABLE_DATA_DIR = MEDIA_ROOT / "table_data"
TABLE_DATA_DIR.mkdir(parents=True, exist_ok=True)


# ────────────────────────────────────────────────
# 공통 유틸
# ────────────────────────────────────────────────
def _log(request: HttpRequest, mode: str, query: str, ok: bool, extra: Dict[str, Any]):
    """간단 서버 로그 + MyLog 테이블에 남기기."""
    try:
        ip = (
            request.META.get("REMOTE_ADDR", "")
            or request.META.get("HTTP_X_FORWARDED_FOR", "")
            or ""
        )
        MyLog.objects.create(
            mode_text=mode[:100],
            query=query[:500],
            ok_flag=ok,
            remote_addr_text=ip[:200],
            extra_json=extra,
        )
    except Exception:
        # 로깅 실패는 서비스에 영향 주지 않도록 무시
        pass


def _safe_media_url(abs_path: str) -> Optional[str]:
    """
    MEDIA_ROOT 내부의 실제 파일 경로 -> 브라우저용 MEDIA_URL 로 변환
    파일이 존재하지 않으면 None
    """
    try:
        p = Path(abs_path).resolve()
        if not p.exists():
            return None
        if str(p).startswith(str(MEDIA_ROOT)):
            rel = p.relative_to(MEDIA_ROOT).as_posix()
            return MEDIA_URL.rstrip("/") + "/" + rel.lstrip("/")
    except Exception:
        pass
    return None


def _int(v, default):
    try:
        return int(v)
    except Exception:
        return default


# ────────────────────────────────────────────────
# (1) 이미지 인덱싱
# ────────────────────────────────────────────────
@never_cache
def media_index_view(request: HttpRequest) -> HttpResponse:
    """
    이미지 업로드 → Vertex 멀티모달 임베딩 → Chroma(media_images) 저장.
    """
    if request.method == "GET":
        return render(
            request,
            "ragapp/media_index.html",
            {
                "allow_upload": PUBLIC_ALLOW_UPLOAD_IMAGES,
                "max_files": PUBLIC_MAX_FILES,
                "max_file_mb": PUBLIC_MAX_FILE_MB,
            },
        )

    # POST (업로드)
    if not PUBLIC_ALLOW_UPLOAD_IMAGES:
        return render(
            request,
            "ragapp/media_index.html",
            {
                "allow_upload": False,
                "error": "이미지 업로드가 비활성화되었습니다.",
            },
        )

    files = request.FILES.getlist("images")[:PUBLIC_MAX_FILES]
    use_caption = bool(request.POST.get("caption_from_name"))

    # ✅ MEDIA_ROOT/images/YYYY/MM 형태로 저장 (중복 'uploads' 방지)
    root = MEDIA_ROOT / "images" / timezone.now().strftime("%Y/%m")
    root.mkdir(parents=True, exist_ok=True)

    cards: List[Dict[str, Any]] = []
    ok, fail = 0, 0

    for f in files:
        status = "OK"
        msg = ""
        pid = "-"
        url = "-"
        mime = "-"
        sha16 = "-"
        try:
            if (f.size or 0) > PUBLIC_MAX_FILE_MB * 1024 * 1024:
                raise RuntimeError(f"{PUBLIC_MAX_FILE_MB}MB 제한 초과")
            safe_name = os.path.basename(f.name)
            ts = timezone.now().strftime("%Y%m%d%H%M%S%f")
            dst = root / f"{ts}_{safe_name}"
            with open(dst, "wb") as out:
                for chunk in f.chunks():
                    out.write(chunk)
            mime = mimetypes.guess_type(str(dst))[0] or "application/octet-stream"

            # 🔹 Vertex MultiModalEmbeddingModel 로 이미지 임베딩
            vec = embed_image_file(str(dst), mime=mime)
            pid = add_image_item(
                path=str(dst),
                embedding=vec,
                caption=(dst.stem if use_caption else ""),
            )

            # SHA-256 축약값 (중복 체크/디버깅용)
            h = hashlib.sha256()
            with open(dst, "rb") as rf:
                for c in iter(lambda: rf.read(8192), b""):
                    h.update(c)
            sha16 = h.hexdigest()[:16]

            # ✅ 존재 파일만 URL 생성
            url = _safe_media_url(str(dst)) or "(비공개 경로)"
            ok += 1
        except Exception as e:
            status = "FAIL"
            msg = str(e)
            fail += 1

        cards.append(
            {
                "status": status,
                "msg": msg,
                "pid": pid,
                "url": url,
                "mime": mime,
                "sha16": sha16,
            }
        )

    _log(
        request,
        "media_index",
        f"{len(files)} files",
        True,
        {"ok": ok, "fail": fail},
    )
    return render(
        request,
        "ragapp/media_index.html",
        {
            "allow_upload": PUBLIC_ALLOW_UPLOAD_IMAGES,
            "max_files": PUBLIC_MAX_FILES,
            "max_file_mb": PUBLIC_MAX_FILE_MB,
            "cards": cards,
            "ok": ok,
            "fail": fail,
        },
    )


# ────────────────────────────────────────────────
# (2) 텍스트→이미지 검색
# ────────────────────────────────────────────────
@never_cache
def media_search_view(request: HttpRequest) -> HttpResponse:
    """
    텍스트(설명) → Vertex 멀티모달 텍스트 임베딩 → Chroma(media_images) 검색.
    """
    q = (request.GET.get("q") or "").strip()
    size = max(1, min(_int(request.GET.get("size"), 12), 48))
    page = max(1, _int(request.GET.get("page"), 1))
    k = max(1, min(_int(request.GET.get("k"), 120), 600))
    hits: List[Dict[str, Any]] = []
    total_considered = 0

    if q:
        try:
            qv = embed_text_mm(q)
            top_n = min(page * size, k)  # 요청 페이지까지 확보
            res = search_images_by_text_embedding(
                text_embedding=qv, k=top_n
            ) or {}
            ids = (res.get("ids") or [[]])[0]
            metas = (res.get("metadatas") or [[]])[0]
            docs = (res.get("documents") or [[]])[0]
            total_considered = len(ids)

            start = (page - 1) * size
            end = min(start + size, len(ids))
            for pid, meta, doc in zip(
                ids[start:end],
                metas[start:end],
                docs[start:end],
            ):
                path = (meta or {}).get("path", "") or (meta or {}).get(
                    "filepath", ""
                )
                url = (meta or {}).get("url") or ""
                if not url and path:
                    url = _safe_media_url(path) or ""

                hits.append(
                    {
                        "pid": pid,
                        "caption": (
                            doc
                            or (meta or {}).get("caption")
                            or "(캡션 없음)"
                        ),
                        "path": path,
                        "url": url,
                    }
                )
        except Exception as e:
            return render(
                request,
                "ragapp/media_search.html",
                {
                    "q": q,
                    "size": size,
                    "page": page,
                    "k": k,
                    "error": str(e),
                },
            )

    has_prev = page > 1
    has_next = (page * size) < min(k, total_considered)
    return render(
        request,
        "ragapp/media_search.html",
        {
            "q": q,
            "size": size,
            "page": page,
            "k": k,
            "hits": hits,
            "has_prev": has_prev,
            "has_next": has_next,
        },
    )


# ────────────────────────────────────────────────
# (3) 표 인덱싱 (CSV + 엑셀)
#   - 업로드한 표를 JSON(원본) + Vertex 임베딩 + Chroma(table_rows)에 함께 저장
# ────────────────────────────────────────────────
@never_cache
def table_index_view(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": PUBLIC_ALLOW_UPLOAD_CSV,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
            },
        )

    if not PUBLIC_ALLOW_UPLOAD_CSV:
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": False,
                "error": "표 업로드가 비활성화되어 있습니다.",
            },
        )

    table_name = (request.POST.get("table_name") or "").strip()
    f = request.FILES.get("csvfile")
    limit = _int(request.POST.get("limit"), 0)

    if not table_name or not f:
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": PUBLIC_ALLOW_UPLOAD_CSV,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "error": "표 이름과 파일을 모두 입력해 주세요.",
            },
        )

    # ✅ MEDIA_ROOT/tables/YYYY/MM 에 저장
    root = MEDIA_ROOT / "tables" / timezone.now().strftime("%Y/%m")
    root.mkdir(parents=True, exist_ok=True)
    safe_name = os.path.basename(f.name)
    dst = root / f"{timezone.now().strftime('%Y%m%d%H%M%S%f')}_{safe_name}"

    rows: List[Dict[str, Any]] = []

    try:
        # 1) 파일 저장
        with open(dst, "wb") as out:
            for chunk in f.chunks():
                out.write(chunk)

        # 2) 확장자에 따라 CSV / 엑셀 분기
        ext = dst.suffix.lower()

        if ext in [".xlsx", ".xls"]:
            # 엑셀 처리
            try:
                import openpyxl  # 필요하면: pip install openpyxl
            except ImportError:
                raise RuntimeError(
                    "엑셀 파일(.xlsx, .xls)을 읽으려면 'openpyxl' 패키지가 필요합니다. "
                    "터미널에서 'pip install openpyxl' 실행 후 다시 시도해 주세요."
                )

            wb = openpyxl.load_workbook(dst, data_only=True)
            sheet = wb.active

            # 첫 행을 헤더로 사용
            header = None
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [
                    str(c).strip() if c is not None else "" for c in row
                ]
                break

            if not header or all(not h for h in header):
                raise RuntimeError(
                    "엑셀 파일에서 열 이름(첫 줄)을 찾을 수 없습니다."
                )

            max_rows = PUBLIC_MAX_CSV_ROWS
            limit_rows = min(limit or max_rows, max_rows)

            for idx, row in enumerate(
                sheet.iter_rows(min_row=2, values_only=True),
                start=2,
            ):
                rec: Dict[str, Any] = {}
                for i, col in enumerate(header):
                    col_name = col or f"col_{i+1}"
                    val = row[i] if row and i < len(row) else None
                    rec[col_name] = val
                rows.append(rec)
                if limit_rows and len(rows) >= limit_rows:
                    break

        else:
            # CSV 처리 (기본값)
            with open(dst, "r", encoding="utf-8", newline="") as rf:
                reader = csv.DictReader(rf)
                max_rows = PUBLIC_MAX_CSV_ROWS
                limit_rows = min(limit or max_rows, max_rows)
                for row in reader:
                    rows.append(row)
                    if limit_rows and len(rows) >= limit_rows:
                        break

        # ⚠️ 행이 하나도 없으면 바로 리턴
        if not rows:
            return render(
                request,
                "ragapp/table_index.html",
                {
                    "allow_upload": PUBLIC_ALLOW_UPLOAD_CSV,
                    "max_rows": PUBLIC_MAX_CSV_ROWS,
                    "error": "표 안에 읽을 수 있는 줄이 없습니다.",
                },
            )

        # 3) 스키마 / 샘플 정보 DB에 저장 (TableSchema)
        try:
            if TableSchema is not None:
                cols = list(rows[0].keys())
                sample_rows = rows[:5]

                from datetime import datetime

                def _infer_type(val: Any) -> str:
                    if val is None:
                        return "text"
                    s = str(val).strip()
                    if not s:
                        return "text"
                    # 숫자처럼 보이면 number
                    try:
                        float(str(s).replace(",", ""))
                        return "number"
                    except Exception:
                        pass
                    # ISO 형식 날짜/시간 추정
                    try:
                        datetime.fromisoformat(s)
                        return "date"
                    except Exception:
                        return "text"

                column_types: Dict[str, str] = {}
                for col in cols:
                    inferred = "text"
                    for r in sample_rows:
                        v = r.get(col)
                        if v not in (None, ""):
                            inferred = _infer_type(v)
                            break
                    column_types[col] = inferred

                TableSchema.objects.update_or_create(
                    table_name=table_name,
                    defaults={
                        "columns": cols,
                        "column_types": column_types,
                        "sample_rows": sample_rows,
                    },
                )
        except Exception:
            # 스키마 저장에 실패해도 업로드/인덱싱은 계속 간다
            pass

        # 4) 원본 행 전체를 JSON 파일로도 보관 (LLM 필터 + 키워드 fallback용)
        try:
            json_path = TABLE_DATA_DIR / f"{table_name}.json"
            with open(json_path, "w", encoding="utf-8") as jf:
                json.dump(rows, jf, ensure_ascii=False)
        except Exception:
            # JSON 저장 실패해도 인덱싱은 계속
            pass

        # 5) 벡터 인덱싱 (Vertex text-embedding-004 + Chroma)
        def row_to_str(r: Dict[str, Any]) -> str:
            return " | ".join(f"{k}:{r.get(k,'')}" for k in r.keys())

        texts = [row_to_str(r) for r in rows]
        if not texts:
            raise RuntimeError("인덱싱할 행이 없습니다.")

        embs = embed_texts(texts)
        added = add_table_rows(
            table_name=table_name,
            rows=rows,
            embeddings=embs,
        )

        _log(
            request,
            "table_index",
            table_name,
            True,
            {"rows": len(rows), "added": added},
        )

        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": PUBLIC_ALLOW_UPLOAD_CSV,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "ok": True,
                "table_name": table_name,
                "added": added,
                "filename": dst.name,
            },
        )

    except Exception as e:
        rows_count = len(rows) if isinstance(rows, list) else 0
        _log(
            request,
            "table_index",
            table_name,
            False,
            {"rows": rows_count, "error": str(e)},
        )
        return render(
            request,
            "ragapp/table_index.html",
            {
                "allow_upload": PUBLIC_ALLOW_UPLOAD_CSV,
                "max_rows": PUBLIC_MAX_CSV_ROWS,
                "error": str(e),
            },
        )


# ────────────────────────────────────────────────
# (4) 표 검색 + 그룹/집계 (자동 추론 + Vertex LLM/임베딩)
# ────────────────────────────────────────────────

ALLOWED_AGG = {"", "count", "sum", "avg", "min", "max"}

# 질문에서 집계 의도를 대충 읽어내기 위한 힌트 (기본값)
AGG_HINTS: dict[str, list[str]] = {
    "sum": ["합계", "총 ", "전체", "총액", "총매출", "total"],
    "avg": ["평균", "평균적으로", "average", "avg"],
    "max": ["최대", "가장 큰", "제일 큰", "가장 높은", "top"],
    "min": ["최소", "가장 작은", "가장 낮은"],
    "count": ["개수", "건수", "몇 개", "몇개", "몇 명", "몇명", "row 수"],
}

# 컬럼 의미에 대한 한국어/영어 별칭 (기본값)
COLUMN_SYNONYMS: dict[str, list[str]] = {
    "region": ["지역", "지역별", "도시", "시도", "branch", "지점"],
    "product": ["상품", "메뉴", "제품", "메뉴명", "item"],
    "channel": ["채널", "판매채널", "판매 경로", "sales channel"],
    "date": ["날짜", "일자", "date", "일별"],
    "sales": ["매출", "매출액", "금액", "판매금액", "revenue", "sales"],
}

NUMERIC_HINTS = ["sales", "amount", "revenue", "price", "qty", "quantity", "count"]


def _safe_json_dict(text: Any) -> Dict[str, Any]:
    if not text:
        return {}
    try:
        data = json.loads(text)
        if isinstance(data, dict):
            return data
    except Exception:
        pass
    return {}


def _safe_json_list(text: Any) -> List[Any]:
    if not text:
        return []
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except Exception:
        pass
    return []


def _load_table_search_config(
    table: str,
) -> Tuple[Dict[str, List[str]], Dict[str, List[str]], List[str], float, bool]:
    """
    TableSearchRule 에서 현재 검색에 쓸 규칙을 로드.
    - table 에 맞는 활성 규칙이 있으면 우선 사용
    - 없으면 table_name 이 비어있는 '공통 규칙'을 사용
    - 결국 못 찾으면 코드 기본값으로 반환
    """
    agg_hints = {k: list(v) for k, v in AGG_HINTS.items()}
    column_synonyms = {k: list(v) for k, v in COLUMN_SYNONYMS.items()}
    numeric_hints = list(NUMERIC_HINTS)
    min_sim = 0.35
    hard_filter_enabled = True

    if TableSearchRule is None:
        return agg_hints, column_synonyms, numeric_hints, min_sim, hard_filter_enabled

    try:
        qs = TableSearchRule.objects.filter(is_active=True)
        rule = None
        if table:
            rule = (
                qs.filter(table_name=table)
                .order_by("-updated_at", "-id")
                .first()
            )
        if rule is None:
            rule = (
                qs.filter(table_name__in=["", None])
                .order_by("-updated_at", "-id")
                .first()
            )
    except Exception:
        rule = None

    if not rule:
        return agg_hints, column_synonyms, numeric_hints, min_sim, hard_filter_enabled

    # min_sim
    try:
        if getattr(rule, "min_sim", None) is not None:
            min_sim = float(rule.min_sim)
    except Exception:
        pass

    hard_filter_enabled = bool(getattr(rule, "hard_filter_enabled", True))

    # agg_hints_json: {"sum":["합계","총액"], ...}
    override_agg = _safe_json_dict(getattr(rule, "agg_hints_json", None))
    for key, words in override_agg.items():
        if isinstance(words, list):
            agg_hints[str(key)] = [str(w) for w in words]
        elif isinstance(words, str):
            agg_hints[str(key)] = [words]

    # column_synonyms_json: {"region":["지역","지점"], ...}
    override_syn = _safe_json_dict(getattr(rule, "column_synonyms_json", None))
    for key, syns in override_syn.items():
        if isinstance(syns, list):
            column_synonyms[str(key)] = [str(s) for s in syns]
        elif isinstance(syns, str):
            column_synonyms[str(key)] = [syns]

    # numeric_hints_json: ["sales","amount", ...]
    override_num = _safe_json_list(getattr(rule, "numeric_hints_json", None))
    if override_num:
        numeric_hints = [str(x) for x in override_num]

    return agg_hints, column_synonyms, numeric_hints, min_sim, hard_filter_enabled


def _get_table_schema_info(table_name: str) -> tuple[list[str], dict[str, str]]:
    """
    TableSchema 에서 컬럼 이름 리스트와 column_types(dict)를 최대한 안전하게 꺼냄.
    """
    if TableSchema is None or not table_name:
        return [], {}

    try:
        schema = (
            TableSchema.objects.filter(table_name=table_name)
            .order_by("-updated_at", "-created_at", "-id")
            .first()
        )
    except Exception:
        schema = None

    if schema is None:
        return [], {}

    cols_raw = getattr(schema, "columns", None)
    col_types_raw = getattr(schema, "column_types", None) or getattr(
        schema, "column_types_json", None
    )

    # columns 파싱
    cols: list[str] = []
    if isinstance(cols_raw, list):
        for item in cols_raw:
            if isinstance(item, dict):
                name = item.get("name") or item.get("column") or item.get("key")
            else:
                name = str(item)
            if name and name not in cols:
                cols.append(name)
    elif isinstance(cols_raw, dict):
        cols = list(cols_raw.keys())
    elif isinstance(cols_raw, str):
        try:
            j = json.loads(cols_raw)
            if isinstance(j, list):
                for item in j:
                    if isinstance(item, dict):
                        name = item.get("name") or item.get("column") or item.get("key")
                    else:
                        name = str(item)
                    if name and name not in cols:
                        cols.append(name)
            elif isinstance(j, dict):
                cols = list(j.keys())
        except Exception:
            cols = [c.strip() for c in cols_raw.split(",") if c.strip()]

    # column_types 파싱
    col_types: dict[str, str] = {}
    if isinstance(col_types_raw, dict):
        col_types = {str(k): str(v) for k, v in col_types_raw.items()}
    elif isinstance(col_types_raw, str):
        try:
            j = json.loads(col_types_raw)
            if isinstance(j, dict):
                col_types = {str(k): str(v) for k, v in j.items()}
        except Exception:
            pass

    return cols, col_types


def _guess_agg_from_question(q: str, agg_hints: Dict[str, List[str]]) -> str:
    q_lower = q.lower()
    for agg_key, words in agg_hints.items():
        for w in words:
            if w in q or w in q_lower:
                return agg_key
    return ""


def _auto_fill_table_and_agg(
    q: str,
    table: str,
    group_by: str,
    agg_field: str,
    agg: str,
    agg_hints: Dict[str, List[str]],
    column_synonyms: Dict[str, List[str]],
    numeric_hints: List[str],
) -> tuple[str, str, str, str]:
    """
    사용자가 고급 설정을 비웠을 때, TableSchema + 질문을 보고
    table / group_by / agg_field / agg 를 최대한 자동으로 채워줌.
    이미 사용자가 적은 값은 건드리지 않음.
    """
    # 0) table 비어 있고, TableSchema 에 표가 딱 1개면 자동 선택
    if (not table) and TableSchema is not None:
        try:
            distinct_tables = list(
                TableSchema.objects.order_by("table_name")
                .values_list("table_name", flat=True)
                .distinct()
            )
        except Exception:
            distinct_tables = []
        if len(distinct_tables) == 1:
            table = distinct_tables[0]

    # TableSchema 없으면 여기까지
    if TableSchema is None or not table:
        return table, group_by, agg_field, agg

    cols, col_types = _get_table_schema_info(table)
    if not cols:
        return table, group_by, agg_field, agg

    # 1) agg 비어 있으면 질문에서 추정
    if not agg:
        agg = _guess_agg_from_question(q, agg_hints)

    # 2) agg_field 비어 있으면 숫자 컬럼 중에서 추정
    if not agg_field and agg:
        numeric_cols = [
            c
            for c in cols
            if col_types.get(c) in ("number", "numeric", "float", "int")
        ]
        chosen = ""
        q_lower = q.lower()

        for c in numeric_cols:
            name_lower = c.lower()
            if any(h in name_lower for h in numeric_hints):
                chosen = c
                break
            if name_lower in q_lower:
                chosen = c
                break

        if not chosen and len(numeric_cols) == 1:
            chosen = numeric_cols[0]

        if chosen:
            agg_field = chosen

    # 3) group_by 비어 있으면 텍스트/카테고리 컬럼에서 추정
    if not group_by and agg and agg_field:
        q_lower = q.lower()
        text_cols = [
            c
            for c in cols
            if col_types.get(c, "text") not in ("number", "numeric", "float", "int")
        ]

        candidate_scores: list[tuple[int, str]] = []

        for c in text_cols:
            score = 0
            name_lower = c.lower()

            if name_lower in q_lower:
                score += 5

            for key, syns in column_synonyms.items():
                if key in name_lower:
                    if any(s in q for s in syns):
                        score += 4
                    else:
                        score += 1

            if score > 0:
                candidate_scores.append((score, c))

        if candidate_scores:
            candidate_scores.sort(reverse=True)
            group_by = candidate_scores[0][1]
        else:
            if len(text_cols) == 1:
                group_by = text_cols[0]

    return table, group_by, agg_field, agg


def _infer_columns(hit_rows: List[Dict[str, Any]], table: Optional[str]) -> List[str]:
    """
    검색된 행들에서 컬럼 이름을 추려서 보여줄 순서를 정함.
    - table 이 None 이면 여러 표가 섞인 상황 → _table 컬럼을 맨 앞으로 보냄.
    """
    if not hit_rows:
        return []

    seen: List[str] = []
    for r in hit_rows:
        if not isinstance(r, dict):
            continue
        for k in r.keys():
            if k not in seen:
                seen.append(k)

    if not table and "_table" in seen:
        seen.remove("_table")
        seen.insert(0, "_table")

    return seen


def _to_float(v: Any) -> Optional[float]:
    """집계용 숫자 변환 (문자열/콤마 포함도 처리). 실패하면 None."""
    try:
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(",", "").strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _apply_group_agg(
    hit_rows: List[Dict[str, Any]],
    group_by: str,
    agg: str,
    agg_field: str,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    group_by / agg / agg_field 설정에 따라 그룹별 count/sum/avg/min/max 계산.
    """
    groups: Dict[str, List[float]] = {}
    counts: Dict[str, int] = {}

    for r in hit_rows:
        if not isinstance(r, dict):
            continue

        key = str(r.get(group_by, "") or "(값 없음)")
        counts[key] = counts.get(key, 0) + 1

        if agg == "count" or not agg_field:
            continue

        val = _to_float(r.get(agg_field))
        if val is None:
            continue
        groups.setdefault(key, []).append(val)

    rows_out: List[Dict[str, Any]] = []

    if agg == "count" or not agg_field:
        columns = [group_by, "rows"]
        for key, count in counts.items():
            rows_out.append({group_by: key, "rows": count})
    else:
        col_name = f"{agg}_{agg_field}"
        columns = [group_by, "rows", col_name]
        for key, nums in groups.items():
            if not nums:
                continue

            if agg == "sum":
                value = sum(nums)
            elif agg == "avg":
                value = sum(nums) / len(nums)
            elif agg == "min":
                value = min(nums)
            elif agg == "max":
                value = max(nums)
            else:
                value = len(nums)

            rows_out.append(
                {
                    group_by: key,
                    "rows": counts.get(key, len(nums)),
                    col_name: value,
                }
            )

    rows_out.sort(key=lambda r: str(r.get(group_by, "")))
    return rows_out, columns


def _load_table_rows_from_file(table: str) -> List[Dict[str, Any]]:
    """
    TABLE_DATA_DIR/table_name.json 에서 원본 행 전체를 로드.
    파일이 없거나 깨져 있으면 빈 리스트.
    """
    try:
        if not table:
            return []
        path = TABLE_DATA_DIR / f"{table}.json"
        if not path.exists():
            return []
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return [r for r in data if isinstance(r, dict)]
        return []
    except Exception:
        return []


def _apply_filters(
    rows: List[Dict[str, Any]],
    filters: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """
    LLM 이 만들어준 filters 를 그대로 적용.
    필터 형식 예:
      {"column": "region", "op": "=", "value": "서울"}
      {"column": "product", "op": "in", "value": ["아메리카노", "라떼"]}
      {"column": "sales", "op": ">", "value": 1000000}
    """
    if not filters:
        return rows

    def _match(row: Dict[str, Any], flt: Dict[str, Any]) -> bool:
        col = flt.get("column") or flt.get("field")
        if not col:
            return True
        op = (flt.get("op") or "=").lower()
        val = flt.get("value")
        cell = row.get(col)
        if cell is None:
            return False

        s = str(cell)
        if op in ("=", "eq"):
            if isinstance(val, list):
                return s in [str(v) for v in val]
            return s == str(val)
        if op in ("contains", "like"):
            return str(val) in s
        if op == "in":
            if not isinstance(val, list):
                return False
            return s in [str(v) for v in val]

        # 숫자 비교 (>, >=, <, <=)
        try:
            cnum = float(str(cell).replace(",", ""))
            vnum = float(str(val).replace(",", ""))
        except Exception:
            return False

        if op in (">", "gt"):
            return cnum > vnum
        if op in (">=", "ge"):
            return cnum >= vnum
        if op in ("<", "lt"):
            return cnum < vnum
        if op in ("<=", "le"):
            return cnum <= vnum

        return True

    out: List[Dict[str, Any]] = []
    for r in rows:
        ok = True
        for f in filters:
            if not _match(r, f):
                ok = False
                break
        if ok:
            out.append(r)
    return out


def _hard_filter_rows_by_question(
    q: str,
    rows: List[Dict[str, Any]],
    columns: List[str],
) -> List[Dict[str, Any]]:
    """
    질문(q)에 들어있는 단어가 실제 행의 값과 정확히/부분적으로 겹치면
    그 값을 가진 행만 우선적으로 남기는 하드 필터.

    - 예: 질문에 "서울", "아메리카노"가 들어있고
      region / product 컬럼에 그런 값이 있으면
      해당 값이 들어간 행만 우선적으로 남김.
    - 필터 결과가 0건이면, 원래 rows를 그대로 돌려줘서 '전부 날아가는' 일은 막는다.
    """
    if not q or not rows or not columns:
        return rows

    q_norm = q.replace(" ", "")
    candidates = rows

    for col in columns:
        if col == "_table":
            continue

        values = sorted(
            {
                str(r.get(col, "")).strip()
                for r in rows
                if r.get(col) not in (None, "")
            }
        )
        if not values:
            continue

        hit_vals = []
        for v in values:
            v_norm = v.replace(" ", "")
            if not v_norm:
                continue
            if v in q or v_norm in q_norm:
                hit_vals.append(v)

        if hit_vals:
            hit_set = set(hit_vals)
            new_candidates = [
                r for r in candidates if str(r.get(col, "")).strip() in hit_set
            ]
            if new_candidates:
                candidates = new_candidates

    return candidates or rows


@require_GET
def table_search_view(request: HttpRequest) -> HttpResponse:
    """
    업로드해 둔 표(table_rows)에서 자연어 질문으로 행/집계 결과를 찾는 뷰.
    - q: 질문/키워드
    - table: 테이블 이름(선택)
    - group_by / agg_field / agg: 그룹·집계 옵션(선택, 비워두면 자동 추론)
      → 자동 추론은 TableSchema(컬럼 이름/타입) + (선택적으로) Vertex LLM 결과를 사용.
    """
    q = (request.GET.get("q") or "").strip()

    def _to_int_param(name: str, default: int) -> int:
        try:
            v = int(request.GET.get(name, default))
            return max(1, v)
        except Exception:
            return default

    size = _to_int_param("size", 12)
    page = _to_int_param("page", 1)
    try:
        k = int(request.GET.get("k", 200))
    except Exception:
        k = 200

    table = (request.GET.get("table") or "").strip()
    group_by = (request.GET.get("group_by") or "").strip()
    agg_field = (request.GET.get("agg_field") or "").strip()
    agg = (request.GET.get("agg") or "").strip().lower()
    if agg not in ALLOWED_AGG:
        agg = ""

    # 🔹 검색 폼 datalist용 테이블 이름 목록
    table_names: list[str] = []
    if TableSchema is not None:
        try:
            table_names = list(
                TableSchema.objects.order_by("table_name")
                .values_list("table_name", flat=True)
                .distinct()
            )
        except Exception:
            table_names = []

    # 🔹 TableSearchRule 에서 규칙값 로드 (없으면 기본값)
    agg_hints_cfg, column_synonyms_cfg, numeric_hints_cfg, min_sim, hard_filter_enabled = _load_table_search_config(table)

    columns: list[str] = []
    rows: list[dict] = []
    total: int = 0
    page_count: int = 1
    error_msg: str | None = None
    used_loose: bool = False  # 너무 느슨한 기준으로 fallback 했는지 표시

    # ❗ 질문 없으면 폼만
    if not q:
        ctx = {
            "q": q,
            "size": size,
            "page": page,
            "k": k,
            "table": table,
            "group_by": group_by,
            "agg_field": agg_field,
            "agg": agg,
            "columns": columns,
            "rows": rows,
            "total": total,
            "page_count": page_count,
            "error_msg": error_msg,
            "table_names": table_names,
            "used_loose": used_loose,
        }
        return render(request, "ragapp/table_search.html", ctx)

    # 1차: TableSchema 기반 자동 채우기
    orig_table = table
    table, group_by, agg_field, agg = _auto_fill_table_and_agg(
        q=q,
        table=table,
        group_by=group_by,
        agg_field=agg_field,
        agg=agg,
        agg_hints=agg_hints_cfg,
        column_synonyms=column_synonyms_cfg,
        numeric_hints=numeric_hints_cfg,
    )

    # table 이 자동으로 채워진 경우, 그 테이블 기준 규칙을 다시 한 번 로드
    if not orig_table and table != orig_table:
        agg_hints_cfg, column_synonyms_cfg, numeric_hints_cfg, min_sim, hard_filter_enabled = _load_table_search_config(table)

    # 2차: (선택) Vertex LLM 으로 질의 구조 해석
    llm_plan: Optional[Dict[str, Any]] = None
    if infer_table_query_with_vertex is not None and TableSchema is not None:
        try:
            # 각 표의 컬럼/타입/샘플 행을 모아서 LLM에 넘길 요약 구성
            tables_for_llm: Dict[str, Dict[str, Any]] = {}
            for tname in table_names:
                cols, col_types = _get_table_schema_info(tname)
                sample_rows: List[Dict[str, Any]] = []
                try:
                    schema_obj = (
                        TableSchema.objects.filter(table_name=tname)
                        .order_by("-updated_at", "-created_at", "-id")
                        .first()
                    )
                    if schema_obj is not None:
                        sr = getattr(schema_obj, "sample_rows", None)
                        if isinstance(sr, list):
                            sample_rows = [
                                r for r in sr if isinstance(r, dict)
                            ]
                        elif isinstance(sr, str):
                            try:
                                j = json.loads(sr)
                                if isinstance(j, list):
                                    sample_rows = [
                                        r for r in j if isinstance(r, dict)
                                    ]
                            except Exception:
                                sample_rows = []
                except Exception:
                    sample_rows = []

                tables_for_llm[tname] = {
                    "columns": cols,
                    "column_types": col_types,
                    "sample_rows": sample_rows,
                }

            llm_plan = infer_table_query_with_vertex(
                question=q,
                tables=tables_for_llm,
                default_table=table or None,
            )
        except Exception as e:
            log.exception("infer_table_query_with_vertex 실패: %s", e)
            llm_plan = None

    # LLM 결과를 기반으로 table / group_by / agg 설정을 보강
    plan_filters: List[Dict[str, Any]] = []
    if isinstance(llm_plan, dict):
        plan_table = (llm_plan.get("table") or "").strip()
        if plan_table and not table:
            table = plan_table

        if not group_by and llm_plan.get("group_by"):
            group_by = str(llm_plan.get("group_by"))

        if not agg_field and llm_plan.get("agg_field"):
            agg_field = str(llm_plan.get("agg_field"))

        if not agg and llm_plan.get("agg"):
            agg_candidate = str(llm_plan.get("agg")).lower()
            if agg_candidate in ALLOWED_AGG:
                agg = agg_candidate

        pf = llm_plan.get("filters") or llm_plan.get("where") or []
        if isinstance(pf, list):
            plan_filters = [f for f in pf if isinstance(f, dict)]

    try:
        # 3) 질문 → Vertex 텍스트 임베딩 (text-embedding-004)
        q_vecs = embed_texts([q]) or []
        if not q_vecs:
            raise RuntimeError("임베딩을 만들 수 없습니다.")
        qv = q_vecs[0]

        # 4) table_rows에서 k개 검색 (Chroma)
        res = search_table_by_text_embedding(text_embedding=qv, k=k) or {}
        metas_raw = res.get("metadatas") or []
        dists_raw = res.get("distances") or []

        # Chroma 응답 flatten
        if isinstance(metas_raw, list) and metas_raw:
            if isinstance(metas_raw[0], list):
                metas = metas_raw[0] or []
            else:
                metas = metas_raw
        else:
            metas = []

        if isinstance(dists_raw, list) and dists_raw:
            if isinstance(dists_raw[0], list):
                dists = dists_raw[0] or []
            else:
                dists = dists_raw
        else:
            dists = []

        # 길이 맞추기
        if len(dists) < len(metas):
            dists = dists + [None] * (len(metas) - len(dists))
        elif len(dists) > len(metas):
            dists = dists[: len(metas)]

        # 🔹 엄격/느슨 모드 모두 저장해 두고, 나중에 fallback
        strict_all: list[dict] = []        # 유사도 기준 통과 (전체)
        strict_filtered: list[dict] = []   # 유사도 기준 통과 + table 필터 적용
        loose_all: list[dict] = []         # 유사도 낮음까지 포함 (전체)
        loose_filtered: list[dict] = []    # 유사도 낮음까지 포함 + table 필터 적용

        MIN_SIM = float(min_sim or 0.0)

        for meta, dist in zip(metas, dists):
            if not isinstance(meta, dict):
                continue

            meta_table = (meta.get("table") or meta.get("table_name") or "").strip()
            row_json = (
                meta.get("row_json")
                or meta.get("row")
                or meta.get("data")
                or {}
            )

            if isinstance(row_json, str):
                try:
                    row = json.loads(row_json)
                except Exception:
                    row = {}
            elif isinstance(row_json, dict):
                row = row_json
            else:
                row = {}

            if not isinstance(row, dict) or not row:
                continue

            row_with_table = dict(row)
            if meta_table:
                row_with_table["_table"] = meta_table

            # 거리 → 유사도
            try:
                score = 1.0 - float(dist) if dist is not None else None
            except Exception:
                score = None

            match_table = (not table) or (not meta_table) or (meta_table == table)

            # 느슨 모드에는 일단 다 넣기
            loose_all.append(row_with_table)
            if match_table:
                loose_filtered.append(row_with_table)

            # 엄격 모드는 MIN_SIM 이상일 때만
            if (score is None) or (score >= MIN_SIM):
                strict_all.append(row_with_table)
                if match_table:
                    strict_filtered.append(row_with_table)

        # 5) 벡터 기반 1차 선택 (없으면 JSON fallback 시도)
        if strict_filtered:
            parsed_rows = strict_filtered
        elif strict_all:
            parsed_rows = strict_all
        elif loose_filtered:
            parsed_rows = loose_filtered
            used_loose = True
        else:
            parsed_rows = loose_all
            used_loose = True

        # 🔻 여기서까지도 아무 행도 없으면 → JSON 원본 기반 fallback
        if not parsed_rows:
            if table:
                all_rows = _load_table_rows_from_file(table)
                if all_rows:
                    # 간단 키워드 fallback: 질문에 나온 단어들이 들어간 행 우선
                    import re as _re

                    def _row_to_str(r: Dict[str, Any]) -> str:
                        return " ".join(f"{k}:{r.get(k,'')}" for k in r.keys())

                    keywords = [w for w in _re.split(r"\s+", q) if w]
                    fallback_hits: List[Dict[str, Any]] = []
                    for r in all_rows:
                        s = _row_to_str(r)
                        if all(kw in s for kw in keywords):
                            fallback_hits.append(r)

                    parsed_rows = fallback_hits or all_rows
                    used_loose = True

            # JSON에서도 못 찾으면 진짜로 데이터 없음
            if not parsed_rows:
                ctx = {
                    "q": q,
                    "size": size,
                    "page": page,
                    "k": k,
                    "table": table,
                    "group_by": group_by,
                    "agg_field": agg_field,
                    "agg": agg,
                    "columns": [],
                    "rows": [],
                    "total": 0,
                    "page_count": 1,
                    "error_msg": None,
                    "table_names": table_names,
                    "used_loose": used_loose,
                }
                return render(request, "ragapp/table_search.html", ctx)

        # 6) LLM이 제안한 filters 가 있으면 한 번 더 필터링
        if plan_filters:
            filtered = _apply_filters(parsed_rows, plan_filters)
            if filtered:  # 전부 날아가면 너무 빡세니까, 남을 때만 채택
                parsed_rows = filtered

        # 7) 컬럼 순서 (TableSchema 우선, 없으면 자동 추론)
        col_order: list[str] | None = None
        if TableSchema is not None and table:
            cols, _ = _get_table_schema_info(table)
            if cols:
                col_order = cols

        if col_order is None:
            col_order = _infer_columns(parsed_rows, table or None)

        # 8) 질문 안의 값으로 하드 필터 한 번 더 (서울/아메리카노 등)
        if hard_filter_enabled:
            parsed_rows = _hard_filter_rows_by_question(q, parsed_rows, col_order or [])

        # 9) 집계 모드인지 판단
        if group_by and agg and agg_field:
            rows_all, columns = _apply_group_agg(
                parsed_rows, group_by, agg, agg_field
            )
        else:
            rows_all = parsed_rows
            columns = col_order or []

        # 10) 페이지네이션
        total = len(rows_all)
        page_count = max(1, math.ceil(total / max(1, size)))
        if page > page_count:
            page = page_count
        start = max(0, (page - 1) * size)
        end = start + size
        rows = rows_all[start:end]

        # 🔻 그래도 최종적으로 아무 행도 없으면, JSON 전체에서라도 몇 줄 뽑아주기
        if total == 0 and table:
            all_rows = _load_table_rows_from_file(table)
            if all_rows:
                total = len(all_rows)
                page_count = max(1, math.ceil(total / max(1, size)))
                if page > page_count:
                    page = page_count
                start = max(0, (page - 1) * size)
                end = start + size
                rows = all_rows[start:end]
                columns = _infer_columns(all_rows, table)
                used_loose = True

    except Exception as e:
        error_msg = f"{e.__class__.__name__}: {e}"

    ctx = {
        "q": q,
        "size": size,
        "page": page,
        "k": k,
        "table": table,
        "group_by": group_by,
        "agg_field": agg_field,
        "agg": agg,
        "columns": columns,
        "rows": rows,
        "total": total,
        "page_count": page_count,
        "error_msg": error_msg,
        "table_names": table_names,
        "used_loose": used_loose,
    }
    return render(request, "ragapp/table_search.html", ctx)
